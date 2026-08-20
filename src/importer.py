from pathlib import Path
from openpyxl import load_workbook
print("Running importer from:", __file__)
print("Importer loaded")

class ExcelImporter:

    def __init__(self, filename):

        self.filename = Path(filename)

    # =====================================================
    # PREVIEW EXCEL
    # =====================================================

    def preview(self):

        if not self.filename.exists():
            print(f"\nExcel file not found: {self.filename}")
            return

        workbook = load_workbook(
            self.filename,
            data_only=True
        )

        sheet = workbook.active

        print("\n" + "=" * 60)
        print("EXCEL PREVIEW")
        print("=" * 60)

        count = 0

        for row in sheet.iter_rows(min_row=4, values_only=True):

            if row[0] is None:
                continue

            print(row)

            count += 1

            if count == 5:
                break

    # =====================================================
    # IMPORT DATA
    # =====================================================

    def import_to_database(self, db):
        print("Filename =", self.filename)
        print("Exists =", self.filename.exists())

        if not self.filename.exists():
            print(f"\nExcel file not found: {self.filename}")
            return

        workbook = load_workbook(
            self.filename,
            data_only=True
        )

        sheet = workbook.active

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=4, values_only=True):

            try:

                if row[0] is None:
                    continue

                draw_number = int(row[0])

                draw_date = row[1].strftime("%Y-%m-%d")

                numbers = [
                    int(row[2]),
                    int(row[3]),
                    int(row[4]),
                    int(row[5]),
                    int(row[6])
                ]

                powerball = int(row[7])

                before = db.get_draw_count()

                db.add_draw(
                    game="POWERBALL",
                    draw_number=draw_number,
                    draw_date=draw_date,
                    numbers=numbers,
                    special=powerball
                )

                after = db.get_draw_count()

                if after > before:
                    imported += 1
                else:
                    skipped += 1

            except Exception as e:

                skipped += 1
                print(f"Skipped draw {row[0]} : {e}")

        print("\n" + "=" * 60)
        print("IMPORT COMPLETE")
        print("=" * 60)
        print(f"Imported : {imported}")
        print(f"Skipped  : {skipped}")
        print(f"Database : {db.get_draw_count()} total draws")
        print("=" * 60)

